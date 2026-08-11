#!/usr/bin/env python3
"""
SkandiaMäklarna – bevakning av nya bostadsannonser per kontor.

Kör:  python skm_scraper.py
      python skm_scraper.py --first-run     (bygg baslinje utan att larma)
      python skm_scraper.py --office Karlstad

Läser config.json, hämtar objektlistan per kontor, jämför mot state.json,
och för varje NY annons hämtas detaljsidan med mäklaruppgifter.
Resultatet skrivs till Excel och (om aktiverat) till Monday.
"""

import argparse
import json
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode, urljoin

import requests
from bs4 import BeautifulSoup

BASE = "https://www.skandiamaklarna.se"
HERE = Path(__file__).parent
CONFIG_FILE = HERE / "config.json"
STATE_FILE = HERE / "state.json"
EXCEL_FILE = HERE / "nya_objekt.xlsx"

HEADERS = {
    # Var ärlig med vem du är. Lägg in en riktig kontaktadress.
    "User-Agent": "CBMediaBot/1.0 (+https://cbmedia.se; kontakt: info@cbmedia.se)",
    "Accept-Language": "sv-SE,sv;q=0.9",
}

session = requests.Session()
session.headers.update(HEADERS)


# ----------------------------------------------------------------------
# Hjälpare
# ----------------------------------------------------------------------

def polite_get(url, tries=3):
    """Hämta en sida med paus och återförsök. Var snäll mot servern."""
    for attempt in range(tries):
        try:
            r = session.get(url, timeout=30)
            if r.status_code == 200:
                time.sleep(random.uniform(1.5, 3.0))
                return r.text
            if r.status_code in (429, 503):
                wait = 20 * (attempt + 1)
                print(f"    {r.status_code} – väntar {wait}s")
                time.sleep(wait)
                continue
            print(f"    HTTP {r.status_code} för {url}")
            return None
        except requests.RequestException as e:
            print(f"    Nätverksfel ({e}) – försök {attempt + 1}/{tries}")
            time.sleep(8)
    return None


def text_of(node):
    return re.sub(r"\s+", " ", node.get_text(" ", strip=True)) if node else ""


def first_number(s):
    """'2 195 000 kr' -> 2195000 ; '86,5 kvm' -> 86.5"""
    if not s:
        return None
    s = s.replace("\xa0", " ")
    m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)", s)
    if not m:
        return None
    raw = m.group(1).replace(" ", "").replace(",", ".")
    try:
        val = float(raw)
        return int(val) if val.is_integer() else val
    except ValueError:
        return None


def load_json(path, default):
    if Path(path).exists():
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ----------------------------------------------------------------------
# Steg 0 – slå upp kontorets interna ID utifrån dess slug
# ----------------------------------------------------------------------

OFFICE_ID_RE = re.compile(r"officeId=([0-9a-f\-]{8,})")


def resolve_office_id(slug, cache):
    """'helsingborg' -> 'a8426-ac925-...'. Cachas i state.json."""
    if slug in cache:
        return cache[slug]

    url = f"{BASE}/hitta-maklare/{slug}/till-salu/"
    print(f"  Slår upp kontors-ID via {url}")
    html = polite_get(url)
    if not html:
        return None

    m = OFFICE_ID_RE.search(html)
    if not m:
        print(f"  Hittade inget officeId för '{slug}' – stämmer slugen?")
        return None

    cache[slug] = m.group(1)
    print(f"  {slug} -> {cache[slug]}")
    return cache[slug]


# ----------------------------------------------------------------------
# Steg 1 – hämta objektlänkar för ett kontor
# ----------------------------------------------------------------------

OBJ_URL_RE = re.compile(r"/hitta-hem/[a-z0-9\-]+/[^\s\"'<>]*?/(\d{6,})/?$")


def listing_urls_for_office(office_id, max_pages=6):
    """Returnerar {objekt_id: absolut_url} för ett kontor, senast inkomna först."""
    found = {}
    for page in range(max_pages):
        params = {
            "officeId": office_id,
            "sort": "latestPublishedDate|desc",
        }
        if page:
            params["pageIndex"] = page
        url = f"{BASE}/hitta-hem/?{urlencode(params)}"
        print(f"  Sida {page + 1}: {url}")

        html = polite_get(url)
        if not html:
            break

        soup = BeautifulSoup(html, "html.parser")
        before = len(found)

        for a in soup.find_all("a", href=True):
            href = a["href"].split("?")[0].split("#")[0]
            m = OBJ_URL_RE.search(href)
            if m:
                found[m.group(1)] = urljoin(BASE, href)

        new_on_page = len(found) - before
        print(f"    {new_on_page} objekt hittade (totalt {len(found)})")
        if new_on_page == 0:
            break

    return found


# ----------------------------------------------------------------------
# Steg 2 – plocka ut detaljer från en objektsida
# ----------------------------------------------------------------------

def parse_listing(url, office_name):
    html = polite_get(url)
    if not html:
        return None

    soup = BeautifulSoup(html, "html.parser")
    page_text = soup.get_text(" ", strip=True)

    data = {
        "objekt_id": (OBJ_URL_RE.search(url.rstrip("/") + "/") or [None, ""])[1]
        if OBJ_URL_RE.search(url.rstrip("/") + "/") else "",
        "kontor": office_name,
        "url": url,
        "hittad": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    # --- Adress och område: og:title ser ut som "Axeltorpsgatan 5, Helsingborg | Bostadsrätt | ..."
    og_title = soup.find("meta", property="og:title")
    title = og_title["content"] if og_title and og_title.get("content") else ""
    parts = [p.strip() for p in title.split("|")]
    addr_city = parts[0] if parts else ""
    data["bostadstyp"] = parts[1] if len(parts) > 1 else ""
    if "," in addr_city:
        data["adress"], data["ort"] = [x.strip() for x in addr_city.split(",", 1)]
    else:
        data["adress"], data["ort"] = addr_city, ""

    # Fallback på H1
    if not data["adress"]:
        data["adress"] = text_of(soup.find("h1"))

    # --- Rum och boarea.
    # Rubriken ser ut som "3 rum 70 kvm" eller "4 rum 53 + 10 kvm",
    # där andra talet är biarea. Att bara leta närmast "kvm" ger biarean.
    # Därför knyts rum och kvm ihop i samma mönster.
    data["rum"] = None
    data["boarea_kvm"] = None
    data["biarea_kvm"] = None

    m = re.search(
        r"(\d+(?:[,.]\d+)?)\s*rum\s+(\d+(?:[,.]\d+)?)"
        r"(?:\s*\+\s*(\d+(?:[,.]\d+)?))?\s*kvm",
        page_text,
    )
    if m:
        data["rum"] = first_number(m.group(1))
        data["boarea_kvm"] = first_number(m.group(2))
        if m.group(3):
            data["biarea_kvm"] = first_number(m.group(3))
    else:
        # Objekt utan rumsangivelse, t.ex. tomter och gårdar
        m = re.search(r"(\d+(?:[,.]\d+)?)(?:\s*\+\s*(\d+(?:[,.]\d+)?))?\s*kvm", page_text)
        if m:
            data["boarea_kvm"] = first_number(m.group(1))
            if m.group(2):
                data["biarea_kvm"] = first_number(m.group(2))
        m = re.search(r"(\d+(?:[,.]\d+)?)\s*rum", page_text)
        if m:
            data["rum"] = first_number(m.group(1))

    # --- Pris.
    # Först med etikett, sedan utan. Många objekt saknar pris helt
    # (REDO och kommande), och då ska fältet förbli tomt.
    data["pris"] = None
    m = re.search(
        r"(?:Högstbjudande|Utgångspris|Accepterat pris|Prisidé|Startpris|Begärt pris|Pris)"
        r"\s*:?\s*([\d\s\xa0]+)\s*kr",
        page_text,
    )
    if m:
        data["pris"] = first_number(m.group(1))
    else:
        # Fallback: första beloppet som inte är en löpande kostnad
        for cand in re.finditer(r"([\d][\d\s\xa0]{5,})\s*kr(?!\s*/)", page_text):
            val = first_number(cand.group(1))
            if val and val >= 100_000:
                data["pris"] = val
                break

    m = re.search(r"Månadsavgift\s*:?\s*([\d\s\xa0]+)\s*kr", page_text)
    data["avgift"] = first_number(m.group(1)) if m else None

    # --- Mäklare: leta upp mailto mot skandiamaklarna.se
    broker_mail = ""
    for a in soup.select('a[href^="mailto:"]'):
        addr = a["href"].replace("mailto:", "").split("?")[0].strip()
        if addr.lower().endswith("@skandiamaklarna.se"):
            broker_mail = addr
            break
    data["maklare_epost"] = broker_mail

    # Namn: gissa ur mejladressen, men föredra länken till personalsidan
    broker_name, broker_slug = "", ""
    for a in soup.select('a[href*="/personal/"]'):
        name = text_of(a)
        if name and len(name.split()) >= 2 and "Läs mer" not in name:
            broker_name = name
            broker_slug = a["href"].rstrip("/").split("/")[-1]
            break
    if not broker_name:
        m = re.search(r"[Aa]nsvarig mäklare\s+([A-ZÅÄÖ][\wÅÄÖåäö\-]+(?:\s+[A-ZÅÄÖ][\wÅÄÖåäö\-]+)+)", page_text)
        if m:
            broker_name = m.group(1)
    if not broker_name and broker_mail:
        broker_name = broker_mail.split("@")[0].replace(".", " ").title()
    data["maklare"] = broker_name
    data["maklare_profil"] = f"{BASE}/personal/{broker_slug}/" if broker_slug else ""

    # Telefon: första tel-länk som ser svensk ut
    phone = ""
    for a in soup.select('a[href^="tel:"]'):
        p = a["href"].replace("tel:", "").strip()
        if len(re.sub(r"\D", "", p)) >= 8:
            phone = text_of(a) or p
            break
    data["maklare_telefon"] = phone

    # --- Huvudbild
    og_img = soup.find("meta", property="og:image")
    data["bild"] = og_img["content"].split("?")[0] if og_img and og_img.get("content") else ""

    # --- Finns redan film? Då är objektet mindre intressant att sälja på.
    data["har_film"] = "Ja" if "video.skm.quedro.com" in html else "Nej"

    # --- Publiceringsdatum.
    # Sajten visar det inte öppet, men det förekommer i strukturerad data på
    # vissa objekt. Hittas inget används tidpunkten då vi först såg annonsen,
    # vilket med timvis avsökning ligger inom en timme från publiceringen.
    published, published_kalla = "", "Först sedd"

    for meta_key in ("article:published_time", "og:published_time"):
        tag = soup.find("meta", property=meta_key) or soup.find("meta", attrs={"name": meta_key})
        if tag and tag.get("content"):
            published, published_kalla = tag["content"][:10], "Från sidan"
            break

    if not published:
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                blob = json.loads(script.string or "{}")
            except (json.JSONDecodeError, TypeError):
                continue
            for item in (blob if isinstance(blob, list) else [blob]):
                if isinstance(item, dict):
                    for key in ("datePublished", "dateCreated", "uploadDate"):
                        if item.get(key):
                            published, published_kalla = str(item[key])[:10], "Från sidan"
                            break
            if published:
                break

    if not published:
        m = re.search(r"[Pp]ublicerad(?:es)?[:\s]+(\d{4}-\d{2}-\d{2})", page_text)
        if m:
            published, published_kalla = m.group(1), "Från sidan"

    data["publicerad"] = published or datetime.now().strftime("%Y-%m-%d")
    data["datumkalla"] = published_kalla

    return data


# ----------------------------------------------------------------------
# Steg 3 – utdata
# ----------------------------------------------------------------------

COLUMNS = [
    ("publicerad", "Publicerad"),
    ("datumkalla", "Datumkälla"),
    ("hittad", "Hittad"),
    ("kontor", "Kontor"),
    ("adress", "Adress"),
    ("ort", "Ort"),
    ("bostadstyp", "Typ"),
    ("rum", "Rum"),
    ("boarea_kvm", "Boarea"),
    ("biarea_kvm", "Biarea"),
    ("pris", "Pris"),
    ("maklare", "Mäklare"),
    ("maklare_epost", "Mäklare e-post"),
    ("maklare_telefon", "Telefon"),
    ("har_film", "Har film"),
    ("url", "Annons"),
    ("maklare_profil", "Mäklarprofil"),
    ("bild", "Bild"),
    ("objekt_id", "Objekt-ID"),
]


def _col_index(key):
    """1-baserat kolumnnummer för ett fält, så indexen inte spricker
    när COLUMNS ändras."""
    return [k for k, _ in COLUMNS].index(key) + 1


def write_excel(rows):
    """Lägger till nya rader sist i arbetsboken, skapar den om den saknas."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    link_cols = [_col_index(k) for k in ("url", "maklare_profil", "bild")]
    price_col = _col_index("pris")
    status_col = len(COLUMNS) + 1

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Nya objekt"
        ws.append([label for _, label in COLUMNS])
        head_fill = PatternFill("solid", fgColor="0B0D10")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="66FCF1", name="Calibri", size=11)
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        widths = {"publicerad": 13, "datumkalla": 13, "hittad": 16, "kontor": 18,
                  "adress": 26, "ort": 16, "bostadstyp": 14, "rum": 7,
                  "boarea_kvm": 9, "pris": 12, "maklare": 20, "maklare_epost": 30,
                  "maklare_telefon": 15, "har_film": 10, "url": 46,
                  "maklare_profil": 40, "bild": 46, "objekt_id": 12}
        for i, (key, _) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(key, 18)
        ws.cell(row=1, column=status_col, value="Status").font = Font(
            bold=True, color="66FCF1", name="Calibri", size=11
        )
        ws.cell(row=1, column=status_col).fill = head_fill

    for row in rows:
        ws.append([row.get(key, "") for key, _ in COLUMNS])
        r = ws.max_row
        for col in link_cols:
            cell = ws.cell(row=r, column=col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
        ws.cell(row=r, column=price_col).number_format = "# ##0 \"kr\""
        ws.cell(row=r, column=status_col, value="Ej kontaktad")

    wb.save(EXCEL_FILE)
    print(f"\nExcel uppdaterad: {EXCEL_FILE} ({len(rows)} nya rader)")


def push_to_monday(row, cfg):
    """Skapar ett item på din Monday-board. Kolumn-ID:n sätts i config.json."""
    token = cfg.get("monday_token") or os.environ.get("MONDAY_TOKEN", "")
    board = cfg.get("monday_board_id", "")
    if not token or not board:
        return False

    colmap = cfg.get("monday_columns", {})
    values = {}
    for field, col_id in colmap.items():
        val = row.get(field, "")
        if val in (None, ""):
            continue
        # Monday prefixar kolumn-ID med typen. Siffror kan heta både
        # 'numbers_' och 'numeric_' beroende på när kolumnen skapades.
        if col_id.startswith("email"):
            values[col_id] = {"email": str(val), "text": str(val)}
        elif col_id.startswith("phone"):
            # Monday vill ha rena siffror. '073-837 74 74' -> '0738377474'
            digits = re.sub(r"[^\d+]", "", str(val))
            values[col_id] = {"phone": digits, "countryShortName": "SE"}
        elif col_id.startswith("link"):
            values[col_id] = {"url": str(val), "text": "Öppna annons"}
        elif col_id.startswith("date"):
            values[col_id] = {"date": str(val)[:10]}
        elif col_id.startswith(("numbers", "numeric")):
            values[col_id] = str(val)
        else:
            values[col_id] = str(val)

    query = """
    mutation ($board: ID!, $name: String!, $vals: JSON!) {
      create_item (board_id: $board, item_name: $name, column_values: $vals) { id }
    }
    """
    payload = {
        "query": query,
        "variables": {
            "board": str(board),
            "name": f"{row.get('adress', 'Okänd adress')} – {row.get('kontor', '')}",
            "vals": json.dumps(values, ensure_ascii=False),
        },
    }
    try:
        r = requests.post(
            "https://api.monday.com/v2",
            json=payload,
            headers={"Authorization": token, "API-Version": "2024-10"},
            timeout=30,
        )
        body = r.json()
        if "errors" in body:
            print(f"    Monday-fel: {body['errors']}")
            return False
        return True
    except Exception as e:
        print(f"    Monday-fel: {e}")
        return False


# ----------------------------------------------------------------------
# Huvudflöde
# ----------------------------------------------------------------------

def run_scan(cfg, offices, state, max_pages, first_run=False, use_excel=True, limit=0):
    """En avsökning av alla kontor. Returnerar listan med nya objekt.
    limit > 0 behandlar bara så många nya objekt totalt — för testkörningar."""
    seen = state.setdefault("seen", {})
    id_cache = state.setdefault("office_ids", {})
    all_new = []

    for name, slug in offices.items():
        print(f"\n=== {name} ===")
        # Ett färdigt GUID får skrivas direkt i config; annars slås slugen upp.
        looks_like_id = bool(re.fullmatch(r"[0-9a-f]{5}(?:-[0-9a-f]{5})+", slug))
        office_id = slug if looks_like_id else resolve_office_id(slug, id_cache)
        if not office_id:
            continue

        urls = listing_urls_for_office(office_id, max_pages=max_pages)
        if not urls:
            print("  Inga objekt hittades – kolla kontors-ID i config.json")
            continue

        known = set(seen.get(name, []))
        fresh = {oid: u for oid, u in urls.items() if oid not in known}
        print(f"  {len(fresh)} nya av {len(urls)} totalt")

        if not first_run:
            for oid, url in fresh.items():
                if limit and len(all_new) >= limit:
                    print(f"  Testläge: stannar vid {limit} objekt")
                    break
                print(f"  Hämtar {url}")
                row = parse_listing(url, name)
                if not row:
                    continue
                row["objekt_id"] = oid
                all_new.append(row)
                print(f"    {row['adress']} – {row.get('maklare') or 'mäklare okänd'}")

        if limit:
            # I testläge markeras inget som sett, så att en riktig
            # baslinje och skarp körning inte påverkas efteråt.
            pass
        else:
            seen[name] = sorted(set(known) | set(urls))

    if not limit:
        save_json(STATE_FILE, state)

    if first_run or not all_new:
        return all_new

    if use_excel:
        write_excel(all_new)

    if cfg.get("monday_board_id"):
        ok = sum(push_to_monday(r, cfg) for r in all_new)
        print(f"Monday: {ok} av {len(all_new)} rader skapade")

    print("\nNya objekt:")
    for r in all_new:
        pris = f"{r['pris']:,} kr".replace(",", " ") if r.get("pris") else "pris saknas"
        print(f"  • {r['adress']}, {r['ort']} – {pris} – {r.get('maklare', '')} <{r.get('maklare_epost', '')}>")

    return all_new


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--first-run", action="store_true",
                    help="Bygg baslinje: spara allt som redan finns utan att behandla det som nytt")
    ap.add_argument("--office", help="Kör bara ett kontor (namnet i config.json)")
    ap.add_argument("--max-pages", type=int, default=6)
    ap.add_argument("--watch", action="store_true",
                    help="Kör om och om igen istället för en gång")
    ap.add_argument("--interval", type=int, default=60,
                    help="Minuter mellan avsökningar i watch-läge (minst 15)")
    ap.add_argument("--no-excel", action="store_true",
                    help="Skriv bara till Monday, hoppa över Excel")
    ap.add_argument("--limit", type=int, default=0,
                    help="Testläge: behandla bara så här många objekt och rör inte state.json")
    args = ap.parse_args()

    if not CONFIG_FILE.exists():
        sys.exit(f"Saknar {CONFIG_FILE}.")

    cfg = load_json(CONFIG_FILE, {})
    offices = cfg.get("offices", {})
    if args.office:
        if args.office not in offices:
            sys.exit(f"Kontoret '{args.office}' finns inte i config.json")
        offices = {args.office: offices[args.office]}

    state = load_json(STATE_FILE, {"seen": {}, "office_ids": {}})

    # --- Engångskörning ---
    if not args.watch:
        found = run_scan(cfg, offices, state, args.max_pages,
                         first_run=args.first_run, use_excel=not args.no_excel,
                         limit=args.limit)
        if args.limit:
            print(f"\nTestkörning klar: {len(found)} objekt behandlade. "
                  "state.json är orörd, så baslinjen påverkas inte.")
        elif args.first_run:
            total = sum(len(v) for v in state["seen"].values())
            print(f"\nBaslinje sparad: {total} objekt. Nästa körning larmar bara om nytillkomna.")
        elif not found:
            print("\nInga nya annonser sedan förra körningen.")
        return

    # --- Watch-läge ---
    interval = max(15, args.interval)   # under 15 min är onödigt och ohyfsat
    if args.interval < 15:
        print(f"Intervallet höjt till {interval} min – tätare än så ger inget.")

    if not state.get("seen"):
        print("Ingen baslinje hittades. Bygger en först så att du slipper alla befintliga objekt.\n")
        run_scan(cfg, offices, state, args.max_pages, first_run=True)
        print(f"\nBaslinje klar: {sum(len(v) for v in state['seen'].values())} objekt.")

    print(f"\nWatch-läge igång. Söker av var {interval}:e minut. Avbryt med Ctrl+C.\n")
    scans = 0
    try:
        while True:
            scans += 1
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            print(f"\n{'─' * 52}\nAvsökning {scans} – {stamp}")
            try:
                found = run_scan(cfg, offices, state, args.max_pages,
                                 use_excel=not args.no_excel)
                if not found:
                    print("  Inget nytt.")
            except Exception as e:
                # En trasig körning ska inte döda bevakningen
                print(f"  Avsökningen misslyckades: {e}")
            print(f"\nNästa avsökning {interval} min. Sover…")
            time.sleep(interval * 60)
    except KeyboardInterrupt:
        print(f"\n\nAvslutat efter {scans} avsökningar.")


if __name__ == "__main__":
    main()
